from .sheet_parser import parse_sheet
from .flatten import flatten_parsed_sheets

from openpyxl import load_workbook
import re


def load_file(file_path):
    """
    Load an Excel file and return a pandas DataFrame. 

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        pd.DataFrame: DataFrame containing the data from the Excel file.
    """

    # Load the workbook
    wb = load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    # Filter sheet names to only include those that match the pattern
    parsed_sheets = []
    for sheet_name in sheet_names:
        if not re.fullmatch(r'F\d+C\d+', sheet_name):
            continue
        try:
            parsed_sheets.append(parse_sheet(wb,sheet_name))
        except Exception as e:
            print(f"Error parsing sheet {sheet_name}: {e}")
            continue

    return flatten_parsed_sheets(parsed_sheets)


